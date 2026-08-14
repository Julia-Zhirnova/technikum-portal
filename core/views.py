from django.shortcuts import render

# Create your views here.



import io
import csv
import logging
from datetime import datetime
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.core.files.storage import default_storage
from django.utils import timezone
from rest_framework import viewsets, status, parsers
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import TokenAuthentication
from rest_framework_jwt.authentication import JSONWebTokenAuthentication
from core.models import (
    Student, Organization, Employment, EmploymentType,
    ImportHistory, AuditLog
)
from core.serializers import EmploymentSerializer
from core.permissions import IsAdmin, IsCurator, IsTeacher
import openpyxl
from openpyxl import load_workbook
import re
from validate_snils import validate_snils
import hashlib

logger = logging.getLogger(__name__)

class EmploymentImportExportViewSet(viewsets.ModelViewSet):
    """ViewSet для импорта и экспорта трудоустройств"""
    serializer_class = EmploymentSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JSONWebTokenAuthentication]
    
    def get_queryset(self):
        """Получение queryset с учетом прав"""
        user = self.request.user
        if user.is_superuser or user.groups.filter(name='Администратор').exists():
            return Employment.objects.all()
        elif user.groups.filter(name='Куратор').exists():
            # Куратор видит только студентов своей группы
            return Employment.objects.filter(
                student__group__in=user.curator_groups.all()
            )
        elif user.groups.filter(name='Преподаватель').exists():
            # Преподаватель видит только студентов своих групп
            return Employment.objects.filter(
                student__group__in=user.teacher_groups.all()
            )
        else:
            return Employment.objects.none()
    
    @action(detail=False, methods=['post'], url_path='import/dry-run')
    def dry_run_import(self, request):
        """Сухой прогон импорта"""
        return self._process_import(request, dry_run=True)
    
    @action(detail=False, methods=['post'], url_path='import')
    def import_employment(self, request):
        """Импорт трудоустройств"""
        return self._process_import(request, dry_run=False)
    
    def _process_import(self, request, dry_run=False):
        """Обработка импорта"""
        # Проверка прав
        if not self._check_permissions(request):
            return Response(
                {'detail': 'У вас нет прав на импорт'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Проверка файла
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response(
                {'detail': 'Файл не предоставлен'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Проверка размера файла (макс 10 МБ)
        if file_obj.size > 10 * 1024 * 1024:
            return Response(
                {'detail': 'Размер файла превышает 10 МБ'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Проверка расширения
        allowed_extensions = ['.xlsx', '.csv', '.txt']
        file_name = file_obj.name.lower()
        if not any(file_name.endswith(ext) for ext in allowed_extensions):
            return Response(
                {'detail': f'Поддерживаются только форматы: {", ".join(allowed_extensions)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Получение режима импорта
        mode = request.data.get('mode', 'skip_errors')
        if mode not in ['skip_errors', 'stop_on_error', 'update_existing']:
            return Response(
                {'detail': f'Неизвестный режим импорта: {mode}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Проверка антивирусом (заглушка для ClamAV)
        # TODO: Реализовать проверку через ClamAV
        if self._check_virus(file_obj):
            return Response(
                {'detail': 'Файл содержит вредоносный код'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Чтение и парсинг файла
        try:
            data_rows = self._parse_file(file_obj)
        except Exception as e:
            return Response(
                {'detail': f'Ошибка при чтении файла: {str(e)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Проверка количества строк (асинхронная обработка >1000)
        if len(data_rows) > 1000 and not dry_run:
            return self._start_async_import(request, data_rows, mode)
        
        # Обработка импорта
        try:
            result = self._process_import_data(request, data_rows, mode, dry_run)
        except Exception as e:
            logger.error(f"Ошибка при импорте: {e}", exc_info=True)
            return Response(
                {'detail': f'Ошибка при импорте: {str(e)}'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return Response(result, status=status.HTTP_200_OK)
    
    def _parse_file(self, file_obj):
        """Парсинг файла"""
        # Определяем формат по расширению
        file_name = file_obj.name.lower()
        
        if file_name.endswith('.xlsx'):
            return self._parse_xlsx(file_obj)
        elif file_name.endswith('.csv'):
            return self._parse_csv(file_obj)
        elif file_name.endswith('.txt'):
            return self._parse_txt(file_obj)
        else:
            raise ValueError('Неподдерживаемый формат файла')
    
    def _parse_xlsx(self, file_obj):
        """Парсинг XLSX файла"""
        wb = load_workbook(file_obj)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and any(row):
                rows.append(row)
        return rows
    
    def _parse_csv(self, file_obj):
        """Парсинг CSV файла"""
        # Поддержка BOM маркера
        content = file_obj.read().decode('utf-8-sig')
        file_obj.seek(0)
        
        # Определение разделителя
        if ';' in content[:100]:
            delimiter = ';'
        else:
            delimiter = ','
        
        # Чтение CSV
        csv_reader = csv.reader(io.StringIO(content), delimiter=delimiter)
        rows = []
        for row in csv_reader:
            if row and any(row):
                rows.append(row)
        return rows
    
    def _parse_txt(self, file_obj):
        """Парсинг TXT файла"""
        content = file_obj.read().decode('utf-8-sig')
        rows = []
        for line in content.split('
')
'):
            if line.strip():
                # Определяем разделитель
                if '	' in line:
                    row = line.split('	')
                elif ';' in line:
                    row = line.split(';')
                else:
                    row = [line]
                if row and any(row):
                    rows.append(row)
        return rows
    
    def _check_virus(self, file_obj):
        """Проверка на вирусы (заглушка)"""
        # Для EICAR теста
        content = file_obj.read()
        file_obj.seek(0)
        if b'X5O!P%@AP[4\PZX54(P^)7CC)7}$EICAR-STANDARD-ANTIVIRUS-TEST-FILE' in content:
            return True
        return False
    
    def _check_permissions(self, request):
        """Проверка прав на импорт"""
        user = request.user
        return (
            user.is_superuser or 
            user.groups.filter(name='Администратор').exists() or
            user.groups.filter(name='Куратор').exists()
        )
    
    @transaction.atomic
    def _process_import_data(self, request, data_rows, mode, dry_run=False):
        """Обработка данных импорта"""
        result = {
            'dry_run': dry_run,
            'total_rows': len(data_rows),
            'valid_rows': 0,
            'invalid_rows': 0,
            'created_rows': 0,
            'updated_rows': 0,
            'skipped_rows': 0,
            'errors': [],
            'warnings': []
        }
        
        # Создание истории импорта
        import_history = None
        if not dry_run:
            import_history = ImportHistory.objects.create(
                user=request.user,
                file_name=request.FILES['file'].name,
                file_size=request.FILES['file'].size,
                mode=mode,
                status='processing',
                total_rows=len(data_rows)
            )
        
        # Построчный разбор и валидация
        row_errors = []
        valid_rows = []
        
        for idx, row in enumerate(data_rows, start=1):
            row_error = self._validate_row(row, idx)
            if row_error:
                row_errors.append(row_error)
                result['invalid_rows'] += 1
                if mode == 'stop_on_error' and not dry_run:
                    raise ValueError(f"Ошибка на строке {idx}: {row_error}")
            else:
                valid_rows.append(row)
                result['valid_rows'] += 1
        
        # Если есть ошибки и режим stop_on_error, то rollback
        if row_errors and mode == 'stop_on_error' and not dry_run:
            raise ValueError(f"Обнаружены ошибки, импорт отменен: {row_errors}")
        
        # Обработка валидных строк
        for row in valid_rows:
            try:
                process_result = self._process_row(row, request.user, mode)
                if process_result['status'] == 'created':
                    result['created_rows'] += 1
                elif process_result['status'] == 'updated':
                    result['updated_rows'] += 1
                elif process_result['status'] == 'skipped':
                    result['skipped_rows'] += 1
                    result['warnings'].append(process_result.get('message', 'Пропущено'))
                result['errors'].extend(process_result.get('errors', []))
            except Exception as e:
                logger.error(f"Ошибка при обработке строки {row}: {e}", exc_info=True)
                if mode == 'stop_on_error' and not dry_run:
                    raise
                result['errors'].append(str(e))
                result['skipped_rows'] += 1
        
        # Обновление истории импорта
        if import_history and not dry_run:
            import_history.status = 'completed'
            import_history.valid_rows = result['valid_rows']
            import_history.invalid_rows = result['invalid_rows']
            import_history.created_rows = result['created_rows']
            import_history.updated_rows = result['updated_rows']
            import_history.skipped_rows = result['skipped_rows']
            import_history.errors = result['errors']
            import_history.warnings = result['warnings']
            import_history.completed_at = timezone.now()
            import_history.save()
            
            # Создание записи в AuditLog
            AuditLog.objects.create(
                user=request.user,
                action_type='import_employment',
                model_name='Employment',
                details={
                    'mode': mode,
                    'created': result['created_rows'],
                    'updated': result['updated_rows'],
                    'skipped': result['skipped_rows'],
                    'errors_count': len(result['errors']),
                    'warnings_count': len(result['warnings'])
                },
                ip_address=self._get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
        
        return result
    
    def _validate_row(self, row, idx):
        """Валидация строки"""
        # TODO: Реализовать полную валидацию всех полей
        # Проверка СНИЛС
        snils = row[0] if len(row) > 0 else ''
        if snils and not self._validate_snils(snils):
            return f"Строка {idx}: Неверная контрольная сумма СНИЛС"
        
        # Проверка ИНН
        inn = row[1] if len(row) > 1 else ''
        if inn and not self._validate_inn(inn):
            return f"Строка {idx}: Неверный формат ИНН"
        
        # Проверка дат
        date_start = row[2] if len(row) > 2 else ''
        date_end = row[3] if len(row) > 3 else ''
        
        if date_start:
            try:
                date_start_dt = datetime.strptime(date_start, '%d.%m.%Y')
                if date_start_dt > datetime.now():
                    return f"Строка {idx}: Дата начала не может быть в будущем"
            except ValueError:
                return f"Строка {idx}: Неверный формат даты начала"
        
        if date_end and date_start:
            try:
                date_start_dt = datetime.strptime(date_start, '%d.%m.%Y')
                date_end_dt = datetime.strptime(date_end, '%d.%m.%Y')
                if date_end_dt < date_start_dt:
                    return f"Строка {idx}: Дата окончания не может быть раньше даты начала"
            except ValueError:
                return f"Строка {idx}: Неверный формат даты окончания"
        
        # Проверка пересечения периодов
        if snils and date_start and date_end:
            # TODO: Проверить пересечения для одного студента
            pass
        
        return None
    
    def _validate_snils(self, snils):
        """Валидация СНИЛС"""
        # Очистка от пробелов и дефисов
        snils_clean = re.sub(r'[-\s]', '', snils)
        if not re.match(r'^\d{11}$', snils_clean):
            return False
        
        # Проверка контрольной суммы
        try:
            return validate_snils(snils_clean)
        except:
            return False
    
    def _validate_inn(self, inn):
        """Валидация ИНН""" (10 или 12 цифр)
        inn_clean = re.sub(r'[-\s]', '', inn)
        return re.match(r'^\d{10}$|^\d{12}$', inn_clean) is not None
    
    def _process_row(self, row, user, mode):
        """Обработка одной строки"""
        # TODO: Реализовать обработку строки
        return {
            'status': 'created',
            'message': 'Строка обработана'
        }
    
    def _get_client_ip(self, request):
        """Получение IP адреса клиента"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip
    
    def _start_async_import(self, request, data_rows, mode):
        """Запуск асинхронного импорта"""
        from core.tasks import async_import_employment
        task = async_import_employment.delay(
            user_id=request.user.id,
            data_rows=data_rows,
            mode=mode,
            file_name=request.FILES['file'].name,
            file_size=request.FILES['file'].size
        )
        
        return Response(
            {
                'task_id': task.id,
                'status': 'processing',
                'message': 'Импорт запущен асинхронно'
            },
            status=status.HTTP_202_ACCEPTED
        )
    
    @action(detail=False, methods=['get'], url_path='export')
    def export_employment(self, request):
        """Экспорт трудоустройств"""
        # Получение queryset с учетом фильтров
        queryset = self.get_queryset()
        
        # Применение фильтров из URL
        # TODO: Реализовать фильтрацию
        
        # Формат экспорта
        format_type = request.query_params.get('format', 'xlsx')
        if format_type not in ['xlsx', 'csv', 'txt']:
            return Response(
                {'detail': 'Неподдерживаемый формат экспорта'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Сериализация данных
        serializer = self.get_serializer(queryset, many=True)
        data = serializer.data
        
        # Экспорт в выбранном формате
        if format_type == 'xlsx':
            return self._export_xlsx(data)
        elif format_type == 'csv':
            return self._export_csv(data)
        elif format_type == 'txt':
            return self._export_txt(data)
    
    def _export_xlsx(self, data):
        """Экспорт в XLSX"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Трудоустройства'
        
        # Заголовки
        headers = ['СНИЛС', 'ИНН организации', 'Дата начала', 'Дата окончания', 'Должность']
        ws.append(headers)
        
        # Данные
        for item in data:
            row = [
                item.get('student', {}).get('snils', ''),
                item.get('organization', {}).get('inn', ''),
                item.get('date_start', ''),
                item.get('date_end', ''),
                item.get('position', '')
            ]
            ws.append(row)
        
        # Сохранение в буфер
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=employment_export.xlsx'
        return response
    
    def _export_csv(self, data):
        """Экспорт в CSV"""
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        
        # Заголовки
        writer.writerow(['СНИЛС', 'ИНН организации', 'Дата начала', 'Дата окончания', 'Должность'])
        
        # Данные
        for item in data:
            row = [
                item.get('student', {}).get('snils', ''),
                item.get('organization', {}).get('inn', ''),
                item.get('date_start', ''),
                item.get('date_end', ''),
                item.get('position', '')
            ]
            writer.writerow(row)
        
        response = HttpResponse(
            output.getvalue().encode('utf-8-sig'),
            content_type='text/csv'
        )
        response['Content-Disposition'] = 'attachment; filename=employment_export.csv'
        return response
    
    def _export_txt(self, data):
        """Экспорт в TXT"""
        output = []
        
        # Заголовки
        output.append('СНИЛС;ИНН организации;Дата начала;Дата окончания;Должность')
        
        # Данные
        for item in data:
            row = f"{item.get('student', {}).get('snils', '')};{item.get('organization', {}).get('inn', '')};{item.get('date_start', '')};{item.get('date_end', '')};{item.get('position', '')}"
            output.append(row)
        
        response = HttpResponse(
            '
'.join(output).encode('utf-8-sig'),
            content_type='text/plain'
        )
        response['Content-Disposition'] = 'attachment; filename=employment_export.txt'
        return response
