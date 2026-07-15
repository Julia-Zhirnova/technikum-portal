from rest_framework import generics, permissions, status
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import Student, Group, Statement, StudentRequest, Notification, StudentPracticePlace, PracticeDiary, PracticeAttestation, Organization
from .serializers_update import StudentProfileUpdateSerializer
from .serializers import (
    StudentRequestSerializer, NotificationSerializer,
    UserProfileSerializer,
    StudentProfileSerializer, GroupWithStudentsSerializer, StatementSerializer,
    StudentPracticePlaceSerializer, PracticeDiarySerializer, PracticeAttestationSerializer
)

# ==============================================================================
# КАСТОМНЫЕ PERMISSIONS
# ==============================================================================



# ==============================================================================
# TEST VIEW ДЛЯ ДИАГНОСТИКИ
# ==============================================================================

class TestExportView(APIView):
    """Простой test view для проверки работы экспорта"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, statement_id):
        from django.http import HttpResponse
        return HttpResponse(f"Test export for statement {statement_id}", content_type='text/plain')

class IsStudent(permissions.BasePermission):
    """Разрешение только для студентов."""
    def has_permission(self, request, view):
        return request.user.is_authenticated and hasattr(request.user, 'student')

class IsCurator(permissions.BasePermission):
    """Разрешение для кураторов и администраторов."""
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        
        # Админы всегда имеют доступ
        if request.user.is_staff:
            return True
        
        # Проверяем роль curator
        from .models import UserRole
        return UserRole.objects.filter(user=request.user, role__id_role='curator').exists()

class IsTeacher(permissions.BasePermission):
    """Разрешение для преподавателей."""
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        
        # Проверяем роль teacher через UserRole
        from .models import UserRole
        return UserRole.objects.filter(user=request.user, role__id_role='teacher').exists()

# ==============================================================================
# API ENDPOINTS ДЛЯ СТУДЕНТА
# ==============================================================================

class StudentProfileView(generics.RetrieveUpdateAPIView):
    """
    GET /api/student/profile/ — получить профиль студента
    PUT/PATCH /api/student/profile/ — обновить профиль студента
    """
    serializer_class = StudentProfileSerializer
    permission_classes = [permissions.IsAuthenticated, IsStudent]
    
    def get_object(self):
        return Student.objects.get(user=self.request.user)

# ==============================================================================
# API ENDPOINTS ДЛЯ КУРАТОРА
# ==============================================================================

class CuratorGroupView(generics.ListAPIView):
    """
    GET /api/curator/group/ — список групп куратора со студентами
    """
    serializer_class = GroupWithStudentsSerializer
    permission_classes = [permissions.IsAuthenticated, IsCurator]
    pagination_class = None
    
    def get_queryset(self):
        user = self.request.user
        # Проверяем, что пользователь имеет роль куратора
        from .models import UserRole
        has_curator_role = UserRole.objects.filter(
            user=user, 
            role__id_role='curator'
        ).exists()
        
        if not has_curator_role:
            return Group.objects.none()
        
        # Возвращаем ТОЛЬКО группы, где пользователь назначен куратором
        return Group.objects.filter(curator=user).order_by('id_group')

class CuratorStudentDetailView(generics.RetrieveAPIView):
    """
    GET /api/curator/students/<snils>/ — анкета конкретного студента
    """
    serializer_class = StudentProfileSerializer
    permission_classes = [permissions.IsAuthenticated, IsCurator]
    lookup_field = 'snils'
    
    def get_queryset(self):
        # Куратор видит только студентов своей группы
        curator_group = Group.objects.filter(curator=self.request.user).first()
        if curator_group:
            return Student.objects.filter(group=curator_group)
        return Student.objects.none()

# ==============================================================================
# API ENDPOINTS ДЛЯ ПРЕПОДАВАТЕЛЯ
# ==============================================================================

class TeacherStatementsView(generics.ListAPIView):
    """
    GET /api/teacher/statements/ — список ведомостей преподавателя
    """
    serializer_class = StatementSerializer
    permission_classes = [permissions.IsAuthenticated, IsTeacher]
    
    def get_queryset(self):
        return Statement.objects.filter(teacher=self.request.user).order_by('-issue_date')

# ==============================================================================
# УНИВЕРСАЛЬНЫЕ ENDPOINTS
# ==============================================================================



class StudentProfileUpdateView(APIView):
    """
    PATCH /api/student/profile/update/ — обновление профиля студента
    """
    permission_classes = [permissions.IsAuthenticated, IsStudent]
    
    def patch(self, request, *args, **kwargs):
        try:
            student = Student.objects.get(user=request.user)
            print(f"🎯 Найден студент для обновления: {student.snils}")
            
            serializer = StudentProfileUpdateSerializer(student, data=request.data, partial=True)
            
            if serializer.is_valid():
                print(f"✅ Данные валидны, вызываем update()")
                updated_student = serializer.update(student, serializer.validated_data)
                print(f"✅ Обновление завершено")
                
                # Возвращаем обновленный профиль
                profile_serializer = StudentProfileSerializer(updated_student)
                return Response(profile_serializer.data)
            else:
                print(f"❌ Ошибки валидации: {serializer.errors}")
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                
        except Student.DoesNotExist:
            return Response({'error': 'Студент не найден'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f"❌ Ошибка: {e}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class WhoAmIView(APIView):
    """
    GET /api/whoami/ — роли текущего пользователя
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        from .models import UserRole
        user = request.user
        roles = []
        
        # Получаем все роли пользователя из таблицы UserRole
        user_roles = UserRole.objects.filter(user=user).select_related('role')
        for ur in user_roles:
            roles.append(ur.role.id_role)
        
        # Добавляем админа, если is_staff
        if user.is_staff and 'admin' not in roles:
            roles.append('admin')
        
        return Response({
            'user_id': user.id_user,
            'email': user.email,
            'full_name': user.get_full_name(),
            'roles': roles,
        })


# ==============================================================================
# API ENDPOINTS ДЛЯ ПОЛЬЗОВАТЕЛЯ (КУРАТОР/ПРЕПОДАВАТЕЛЬ/АДМИН)
# ==============================================================================

class UserProfileView(generics.RetrieveAPIView):
    """
    GET /api/user/profile/ — профиль куратора/преподавателя/админа
    """
    serializer_class = UserProfileSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_object(self):
        return self.request.user


# ==============================================================================
# API ENDPOINTS ДЛЯ ПОЛУЧЕНИЯ СПРАВОЧНИКОВ (ВЫПАДАЮЩИЕ СПИСКИ)
# ==============================================================================

class ReferencesView(APIView):
    """
    GET /api/references/ — получение всех вариантов choices из моделей
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        from .models import Health, Military, Family, FamilyMember, Passport
        
        references = {
            # Здоровье
            'health_status': [{'value': c[0], 'label': c[1]} for c in Health._meta.get_field('status').choices],
            
            # Воинский учет
            'fitness_category': [{'value': c[0], 'label': c[1]} for c in Military._meta.get_field('fitness_category').choices],
            
            # Семья
            'family_status': [{'value': c[0], 'label': c[1]} for c in Family._meta.get_field('status').choices],
            'housing_type': [{'value': c[0], 'label': c[1]} for c in Family._meta.get_field('housing_type').choices],
            
            # Члены семьи
            'relation': [{'value': c[0], 'label': c[1]} for c in FamilyMember._meta.get_field('relation').choices],
            
            # Паспорт
            'gender': [
                {'value': 'мужской', 'label': 'Мужской'},
                {'value': 'женский', 'label': 'Женский'},
            ],
        }
        
        return Response(references)


# ==============================================================================
# API ENDPOINT ДЛЯ ЗАЧЁТКИ СТУДЕНТА
# ==============================================================================











class StudentGradesView(APIView):
    """
    GET /api/student/grades/ — зачётка текущего студента с группировкой по семестрам
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        from .models import StatementGrade, Statement, DisciplineInGroup, DisciplineReference
        
        try:
            student = Student.objects.get(user=request.user)
            
            # Получаем все оценки студента
            grades = StatementGrade.objects.filter(
                student=student
            ).select_related(
                'statement',
                'statement__group',
                'statement__discipline_in_group',
                'statement__discipline_in_group__discipline_ref'
            )
            
            # Группируем по семестрам
            semesters = {}
            all_grades = []
            
            for grade in grades:
                try:
                    dig = grade.statement.discipline_in_group
                    semester = dig.semester if dig and dig.semester else 1  # Если семестр None, используем 1
                    
                    if semester not in semesters:
                        semesters[semester] = []
                    
                    # Получаем название дисциплины из справочника
                    try:
                        ref = dig.discipline_ref
                        discipline_name = ref.name
                    except:
                        discipline_name = 'N/A'
                    
                    grade_data = {
                        'discipline': discipline_name,
                        'grade': grade.grade,
                        'date': grade.date.strftime('%d.%m.%Y') if grade.date else '',
                        'is_retake': grade.is_retake,
                        'statement_number': grade.statement.number,
                        'group': grade.statement.group.id_group,
                    }
                    
                    semesters[semester].append(grade_data)
                    all_grades.append(grade_data)
                    
                except Exception as e:
                    print(f"Ошибка обработки оценки: {e}")
                    import traceback
                    traceback.print_exc()
                    continue
            
            # Подсчитываем статистику
            total_grades = len(all_grades)
            
            # Отличные оценки (5, отлично, зачтено)
            excellent_count = sum(1 for g in all_grades if g['grade'] in ['5', 'отлично', 'Отлично', 'зачтено', 'Зачтено'])
            
            # Хорошие оценки (4 или хорошо)
            good_count = sum(1 for g in all_grades if g['grade'] in ['4', 'хорошо', 'Хорошо'])
            
            # Удовлетворительные (3 или удовлетворительно)
            satisfactory_count = sum(1 for g in all_grades if g['grade'] in ['3', 'удовлетворительно', 'Удовлетворительно'])
            
            # Неудовлетворительные (2, неудовлетворительно, не зачтено)
            unsatisfactory_count = sum(1 for g in all_grades if g['grade'] in ['2', 'неудовлетворительно', 'Неудовлетворительно', 'не зачтено', 'Не зачтено'])
            
            # Пересдачи
            retake_count = sum(1 for g in all_grades if g['is_retake'])
            
            # Средний балл (только числовые оценки 5, 4, 3, 2)
            # "Зачтено" не включаем в средний балл
            numeric_grades = []
            for g in all_grades:
                grade_lower = g['grade'].lower()
                if g['grade'].isdigit():
                    numeric_grades.append(int(g['grade']))
                elif grade_lower == 'отлично':
                    numeric_grades.append(5)
                elif grade_lower == 'хорошо':
                    numeric_grades.append(4)
                elif grade_lower == 'удовлетворительно':
                    numeric_grades.append(3)
                elif grade_lower == 'неудовлетворительно':
                    numeric_grades.append(2)
                # "Зачтено" и "не зачтено" НЕ включаем
            
            average_score = round(sum(numeric_grades) / len(numeric_grades), 2) if numeric_grades else 0
            
            # Формируем ответ
            semesters_list = []
            for semester in sorted(semesters.keys()):
                semesters_list.append({
                    'semester': semester,
                    'grades': semesters[semester]
                })
            
            return Response({
                'student_snils': student.snils,
                'group': student.group.id_group,
                'total_grades': total_grades,
                'excellent_count': excellent_count,
                'good_count': good_count,
                'satisfactory_count': satisfactory_count,
                'unsatisfactory_count': unsatisfactory_count,
                'retake_count': retake_count,
                'average_score': average_score,
                'semesters': semesters_list
            })
            
        except Student.DoesNotExist:
            return Response({'error': 'Студент не найден'}, status=404)
        except Exception as e:
            print(f"Ошибка в StudentGradesView: {e}")
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)




class UpdateGradeView(APIView):
    """
    PATCH /api/teacher/grades/<int:grade_id>/ — изменение оценки преподавателем
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def patch(self, request, grade_id):
        from .models import StatementGrade, Statement
        
        try:
            grade = StatementGrade.objects.get(id_grade=grade_id)
            statement = grade.statement
            
            # Проверяем, что пользователь - преподаватель этой ведомости
            if statement.teacher != request.user:
                return Response(
                    {'error': 'Вы не являетесь преподавателем этой ведомости'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Проверяем статус ведомости
            if statement.status != 'в_работе':
                return Response(
                    {'error': f'Ведомость имеет статус "{statement.status}". Для изменения оценок ведомость должна быть в статусе "в_работе". Используйте API /api/teacher/statements/{{id}}/restore/ для возврата ведомости в работу.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Получаем новую оценку из запроса (поддерживаем оба варианта: "grade" и "оценка")
            new_grade_value = request.data.get('grade') or request.data.get('оценка')
            if new_grade_value is None:
                return Response(
                    {'error': 'Поле "grade" (или "оценка") обязательно. Пример: {"grade": "5"}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Валидация значения оценки - поддерживаем все форматы
            valid_grades = [
                '5', '4', '3', '2',
                '5 (отлично)', '4 (хорошо)', '3 (удовлетворительно)', '2 (неудовлетворительно)',
                'зачтено', 'не_зачтено', 'Зачтено', 'Не зачтено', 'Не зачтено',
                'отлично', 'хорошо', 'удовлетворительно', 'неудовлетворительно',
                'н/а (не допущен)', 'н/а'
            ]
            if new_grade_value not in valid_grades:
                return Response(
                    {'error': f'Недопустимое значение оценки. Допустимые: {", ".join(valid_grades)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Сохраняем старую оценку для логирования
            old_grade_value = grade.grade
            
            # Обновляем оценку
            grade.grade = new_grade_value
            
            # Если есть дата сдачи, обновляем её
            if 'date' in request.data:
                grade.date = request.data['date']
            
            # Если указано, что это пересдача
            if 'is_retake' in request.data:
                grade.is_retake = request.data['is_retake']
            
            grade.save()
            
            return Response({
                'success': True,
                'message': f'Оценка изменена с "{old_grade_value}" на "{new_grade_value}"',
                'grade_id': grade.id_grade,
                'old_grade': old_grade_value,
                'new_grade': new_grade_value
            })
            
        except StatementGrade.DoesNotExist:
            return Response(
                {'error': 'Оценка не найдена'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



class TeacherStatementGradesView(APIView):
    """
    GET /api/teacher/statements/<int:statement_id>/grades/ — оценки по ведомости
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request, statement_id):
        from .models import Statement, StatementGrade
        
        try:
            statement = Statement.objects.get(id_statement=statement_id)
            
            # Проверяем, что пользователь - преподаватель этой ведомости
            if statement.teacher != request.user:
                return Response(
                    {'error': 'Вы не являетесь преподавателем этой ведомости'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Получаем все оценки по ведомости
            grades = StatementGrade.objects.filter(statement=statement).select_related('student')
            
            grades_data = []
            for grade in grades:
                student = grade.student
                grades_data.append({
                    'id_grade': grade.id_grade,
                    'student_snils': student.snils if student else '',
                    'student_name': student.user.get_full_name() if student and student.user else 'N/A',
                    'grade': grade.grade,
                    'date': grade.date.strftime('%d.%m.%Y') if grade.date else '',
                    'is_retake': grade.is_retake,
                })
            
            return Response(grades_data)
            
        except Statement.DoesNotExist:
            return Response(
                {'error': 'Ведомость не найдена'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



# ==============================================================================
# ЭКСПОРТ/ИМПОРТ ОЦЕНОК
# ==============================================================================

class StatementGradesExportView(APIView):
    """
    GET /api/teacher/statements/<id>/export/?format=excel|csv|txt
    Экспорт оценок ведомости в Excel, CSV или TXT
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def dispatch(self, request, *args, **kwargs):
        print("="*80)
        print(f"🚨 [DISPATCH 1] Входящий request.method: {request.method}")
        
        self.args = args
        self.kwargs = kwargs
        
        # DRF оборачивает Django request в свой Request
        request = self.initialize_request(request, *args, **kwargs)
        self.request = request
        
        print(f"🚨 [DISPATCH 2] После initialize_request, request.method: {request.method}")
        print(f"🚨 [DISPATCH 3] request.method.lower(): {request.method.lower()}")
        print(f"🚨 [DISPATCH 4] Метод в http_method_names? {request.method.lower() in self.http_method_names}")
        
        try:
            # DRF проверяет аутентификацию, права и троттлинг
            self.initial(request, *args, **kwargs)
            print(f"🚨 [DISPATCH 5] После initial, request.method: {request.method}")
            print(f"🚨 [DISPATCH 5b] request.user: {request.user}")
            
            if request.method.lower() in self.http_method_names:
                handler = getattr(self, request.method.lower(), self.http_method_not_allowed)
                print(f"🚨 [DISPATCH 6] Выбран handler: {handler.__name__ if hasattr(handler, '__name__') else handler}")
            else:
                handler = self.http_method_not_allowed
                print(f"🚨 [DISPATCH 6] Выбран handler: http_method_not_allowed")
            
            print(f"🚨 [DISPATCH 7] Вызываем handler...")
            response = handler(request, *args, **kwargs)
            print(f"🚨 [DISPATCH 8] Handler вернул статус: {getattr(response, 'status_code', 'NO STATUS')}")
            print("="*80)
            return response
            
        except Exception as exc:
            print(f"🚨 [DISPATCH EXCEPTION] Исключение: {exc}")
            import traceback
            traceback.print_exc()
            print("="*80)
            response = self.handle_exception(exc)
            return response

    def get(self, request, statement_id):
        from .models import Statement, StatementGrade
        from django.http import HttpResponse
        import csv
        import io
        import traceback
        
        print(f"🔍 Export request for statement_id={statement_id}")
        print(f"   User: {request.user}")
        print(f"   Method: {request.method}")
        
        try:
            statement = Statement.objects.get(id_statement=statement_id)
            print(f"✅ Statement found: {statement.number}")
            print(f"   Teacher: {statement.teacher.get_full_name()}")
            print(f"   Request user: {request.user.get_full_name()}")
            print(f"   Match: {statement.teacher == request.user}")
            
            if statement.teacher != request.user:
                print(f"❌ Access denied")
                return Response({'error': 'Нет доступа'}, status=status.HTTP_403_FORBIDDEN)
            
            grades = StatementGrade.objects.filter(statement=statement).select_related('student')
            
            # Поддерживаем как DRF request.query_params, так и Django request.GET
            export_format = getattr(request, 'query_params', request.GET).get('export_format', 'excel').lower()
            
            # Формируем имя файла
            filename_base = f"{statement.number}_{statement.group.id_group}"
            
            if export_format == 'csv':
                # CSV экспорт
                output = io.StringIO()
                writer = csv.writer(output, delimiter=';')
                writer.writerow(['№', 'ФИО', 'СНИЛС', 'Оценка', 'Дата', 'Пересдача'])
                
                for idx, grade in enumerate(grades, 1):
                    writer.writerow([
                        idx,
                        grade.student.user.get_full_name() if grade.student else '',
                        grade.student.snils if grade.student else '',
                        grade.grade or '',
                        grade.date.strftime('%d.%m.%Y') if grade.date else '',
                        'Да' if grade.is_retake else 'Нет'
                    ])
                
                response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8-sig')
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
                return response
                
            elif export_format == 'txt':
                # TXT экспорт
                output = io.StringIO()
                output.write(f"ВЕДОМОСТЬ {statement.number}\n")
                output.write(f"Группа: {statement.group.id_group}\n")
                output.write(f"Дисциплина: {statement.discipline_in_group.discipline_ref.name}\n")
                output.write("=" * 80 + "\n\n")
                
                for idx, grade in enumerate(grades, 1):
                    student_name = grade.student.user.get_full_name() if grade.student else 'N/A'
                    output.write(f"{idx}. {student_name}\n")
                    output.write(f"   СНИЛС: {grade.student.snils if grade.student else 'N/A'}\n")
                    output.write(f"   Оценка: {grade.grade or '—'}\n")
                    output.write(f"   Дата: {grade.date.strftime('%d.%m.%Y') if grade.date else '—'}\n")
                    output.write(f"   Пересдача: {'Да' if grade.is_retake else 'Нет'}\n\n")
                
                response = HttpResponse(output.getvalue(), content_type='text/plain; charset=utf-8')
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.txt"'
                return response
                
            else:
                # Excel экспорт (по умолчанию)
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Оценки"
                
                # Заголовки
                headers = ['№', 'ФИО', 'СНИЛС', 'Оценка', 'Дата', 'Пересдача']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                # Данные
                for idx, grade in enumerate(grades, 2):
                    ws.cell(row=idx, column=1, value=idx - 1)
                    ws.cell(row=idx, column=2, value=grade.student.user.get_full_name() if grade.student else '')
                    ws.cell(row=idx, column=3, value=grade.student.snils if grade.student else '')
                    ws.cell(row=idx, column=4, value=grade.grade or '')
                    ws.cell(row=idx, column=5, value=grade.date.strftime('%d.%m.%Y') if grade.date else '')
                    ws.cell(row=idx, column=6, value='Да' if grade.is_retake else 'Нет')
                
                # Ширина колонок
                ws.column_dimensions['A'].width = 5
                ws.column_dimensions['B'].width = 40
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 25
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 12
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = HttpResponse(
                    output.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
                return response
                
        except Statement.DoesNotExist:
            print(f"❌ Statement {statement_id} not found")
            return Response({'error': 'Ведомость не найдена'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f"❌ Export error: {e}")
            traceback.print_exc()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class StatementGradesImportView(APIView):
    """
    POST /api/teacher/statements/<id>/import/
    Импорт оценок из Excel/CSV
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request, statement_id):
        from .models import Statement, StatementGrade, Student
        from datetime import datetime
        
        try:
            statement = Statement.objects.get(id_statement=statement_id)
            
            if statement.teacher != request.user:
                return Response({'error': 'Нет доступа'}, status=status.HTTP_403_FORBIDDEN)
            
            if statement.status != 'в_работе':
                return Response(
                    {'error': 'Ведомость закрыта. Верните её в работу для импорта.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            file = request.FILES.get('file')
            if not file:
                return Response({'error': 'Файл не загружен'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Читаем файл
            import openpyxl
            import csv
            import io
            
            grades_updated = 0
            grades_created = 0
            errors = []
            
            def process_row(idx, fio, snils, grade_value):
                nonlocal grades_updated, grades_created
                if not fio:
                    return
                
                # Нормализуем СНИЛС: убираем пробелы и тире
                clean_snils = str(snils).replace(' ', '').replace('-', '').strip() if snils else ''
                clean_fio = str(fio).strip()
                clean_grade = str(grade_value).strip() if grade_value else ''
                
                student = None
                # 1. Ищем по СНИЛС (сравниваем очищенные значения)
                if clean_snils:
                    for s in Student.objects.all():
                        if s.snils and s.snils.replace(' ', '').replace('-', '').strip() == clean_snils:
                            student = s
                            break
                
                # 2. Если не нашли по СНИЛС, ищем по ФИО (разбиваем на фамилию и имя)
                if not student and clean_fio:
                    parts = clean_fio.split()
                    if len(parts) >= 2:
                        last_name = parts[0]
                        first_name = parts[1]
                        try:
                            student = Student.objects.get(user__last_name__iexact=last_name, user__first_name__iexact=first_name)
                        except Student.DoesNotExist:
                            pass
                
                if not student:
                    errors.append(f"Строка {idx}: Студент '{clean_fio}' (СНИЛС: {clean_snils}) не найден в системе.")
                    return
                
                grade_obj, created = StatementGrade.objects.get_or_create(
                    statement=statement,
                    student=student,
                    defaults={'grade': clean_grade}
                )
                if not created:
                    if grade_obj.grade != clean_grade:
                        grade_obj.grade = clean_grade
                        grade_obj.save()
                        grades_updated += 1
                else:
                    grades_created += 1

            if file.name.endswith('.xlsx'):
                wb = openpyxl.load_workbook(io.BytesIO(file.read()))
                ws = wb.active
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[1]: # Проверяем наличие ФИО (колонка 2, индекс 1)
                        continue
                    
                    # Ожидаем: №, ФИО, СНИЛС, Оценка, Дата, Пересдача
                    idx = row[0]
                    fio = row[1]
                    snils = row[2]
                    grade_value = row[3]
                    
                    process_row(idx, fio, snils, grade_value)
                    
            elif file.name.endswith('.csv'):
                content = file.read().decode('utf-8-sig')
                reader = csv.reader(io.StringIO(content), delimiter=';')
                next(reader, None)  # Пропускаем заголовок
                
                for row in reader:
                    if not row or len(row) < 3:
                        continue
                    
                    idx = row[0]
                    fio = row[1]
                    snils = row[2]
                    grade_value = row[3] if len(row) > 3 else ''
                    
                    process_row(idx, fio, snils, grade_value)
            
            message = f'Импорт завершён. Создано: {grades_created}, Обновлено: {grades_updated}.'
            if errors:
                message += f' Ошибок: {len(errors)}. Первые 5: ' + '; '.join(errors[:5])
            
            return Response({
                'success': True,
                'message': message,
                'grades_created': grades_created,
                'grades_updated': grades_updated,
                'errors': errors[:10] # Возвращаем первые 10 ошибок для отладки
            })
            
        except Statement.DoesNotExist:
            print(f"❌ Statement {statement_id} not found")
            return Response({'error': 'Ведомость не найдена'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f"❌ Export error: {e}")
            traceback.print_exc()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==============================================================================
# ГЕНЕРАЦИЯ DOCX ДОКУМЕНТОВ
# ==============================================================================

class StatementGenerateDocxView(APIView):
    """
    POST /api/teacher/statements/<id>/generate-docx/?type=zachet|protocol
    Генерация DOCX документа по шаблону
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request, statement_id):
        from .models import Statement, StatementGrade
        from django.conf import settings
        from django.http import HttpResponse
        from docxtpl import DocxTemplate
        import io
        import os
        import re
        
        try:
            statement = Statement.objects.get(id_statement=statement_id)
            
            if statement.teacher != request.user:
                return Response({'error': 'Нет доступа'}, status=status.HTTP_403_FORBIDDEN)
            
            doc_type = request.query_params.get('type', 'zachet')
            
            # Путь к шаблону
            if doc_type == 'protocol':
                template_path = os.path.join(settings.BASE_DIR, 'templates', 'docx', 'протокол_экзамена.docx')
            else:
                template_path = os.path.join(settings.BASE_DIR, 'templates', 'docx', 'зачетная_ведомость.docx')
            
            if not os.path.exists(template_path):
                return Response(
                    {'error': f'Шаблон не найден: {template_path}'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Получаем оценки
            grades = StatementGrade.objects.filter(statement=statement).select_related('student')
            
            # Формируем контекст
            discipline = statement.discipline_in_group.discipline_ref.name if statement.discipline_in_group else ''
            group = statement.group.id_group
            
            # Определяем курс и семестр из discipline_in_group
            semester = statement.discipline_in_group.semester if statement.discipline_in_group else 1
            course = (semester + 1) // 2  # Примерная оценка курса
            
            # Специальность из группы
            specialty = getattr(statement.group, 'specialty', '09.02.07 «Информационные системы и программирование»')
            
            context = {
                'дисциплина': discipline,
                'группа': group,
                'курс': course,
                'семестр': semester,
                'специальность': specialty,
                'преподаватель': request.user.get_full_name(),
                'допущенные': grades.filter(grade__isnull=False).exclude(grade='').count(),
                'не_явилось': grades.filter(grade__isnull=True).count() + 
                             grades.filter(grade='').count(),
                'день': statement.issue_date.strftime('%d') if statement.issue_date else '',
                'месяц': statement.issue_date.strftime('%m') if statement.issue_date else '',
                'год': statement.issue_date.strftime('%Y') if statement.issue_date else '',
            }
            
            # Добавляем ФИО студентов (до 26)
            for i in range(1, 27):
                if i <= grades.count():
                    grade = grades[i-1]
                    context[f'фио{i}'] = grade.student.user.get_full_name() if grade.student else ''
                    if doc_type == 'protocol':
                        context[f'билет{i}'] = i  # Номер билета по порядку
                else:
                    context[f'фио{i}'] = ''
                    if doc_type == 'protocol':
                        context[f'билет{i}'] = ''
            
            # Добавляем оценки студентов
            grades_dict = {}
            for grade in grades:
                if grade.student:
                    student_name = grade.student.user.get_full_name() if grade.student.user else ''
                    grades_dict[student_name] = grade.grade or ''
            
            # Добавляем оценки в контекст (фио1, фио2, ... уже есть, добавляем оценки)
            for i in range(1, 27):
                if i <= len(grades):
                    grade = grades[i-1]
                    student_name = grade.student.user.get_full_name() if grade.student and grade.student.user else ''
                    # Добавляем оценку в отдельное поле
                    context[f'оценка{i}'] = grades_dict.get(student_name, '')
                else:
                    context[f'оценка{i}'] = ''
            
            # Дополнительные поля для протокола
            if doc_type == 'protocol':
                context['фио_неявка'] = ''
                context['час_начала'] = '09'
                context['минуты_начала'] = '00'
                context['час_окончания'] = '12'
                context['минуты_окончания'] = '00'
                context['мнения'] = ''
                context['нарушения'] = ''
            
            # Генерируем документ
            doc = DocxTemplate(template_path)
            doc.render(context)
            
            # Сохраняем в память
            output = io.BytesIO()
            doc.save(output)
            output.seek(0)
            
            # Формируем имя файла
            safe_discipline = re.sub(r'[<>:"/\\|?*.,;()\[\]{}!@#$%^&+=`~]', ' ', discipline)
            safe_discipline = re.sub(r'\s+', ' ', safe_discipline).strip()[:60]
            
            doc_name = 'Протокол' if doc_type == 'protocol' else 'Зачетная_ведомость'
            filename = f"{doc_name}_{safe_discipline}_{group}.docx"
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Statement.DoesNotExist:
            return Response({'error': 'Ведомость не найдена'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==============================================================================
# ПРОСТОЕ ФУНКЦИОНАЛЬНОЕ ПРЕДСТАВЛЕНИЕ ДЛЯ ТЕСТА
# ==============================================================================
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def debug_export_fbv(request, statement_id):
    print(f"!!! 🚀 FBV ВЫЗВАН УСПЕШНО: statement_id={statement_id}")
    print(f"    User: {request.user}")
    return HttpResponse(f"FBV Работает! Statement: {statement_id}", content_type='text/plain')


# ==============================================================================
# ЗАЯВКИ И УВЕДОМЛЕНИЯ СТУДЕНТА
# ==============================================================================

class StudentRequestView(generics.ListCreateAPIView):
    """GET/POST: Список заявок студента и создание новой"""
    serializer_class = StudentRequestSerializer
    permission_classes = [permissions.IsAuthenticated, IsStudent]

    def get_queryset(self):
        return StudentRequest.objects.filter(student=self.request.user.student)

    def perform_create(self, serializer):
        serializer.save(student=self.request.user.student)


class NotificationView(generics.ListAPIView):
    """GET: Список уведомлений текущего пользователя"""
    serializer_class = NotificationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        # Если это студент, возвращаем его уведомления
        if hasattr(user, 'student'):
            return Notification.objects.filter(student=user.student).order_by('-created_at')
        # Для кураторов, преподавателей и админов пока возвращаем пустой QuerySet 
        # (триггеры уведомлений для них будут добавлены позже)
        return Notification.objects.none()


class MarkNotificationReadView(APIView):
    """POST: Отметить уведомление как прочитанное"""
    permission_classes = [permissions.IsAuthenticated, IsStudent]

    def post(self, request, notification_id):
        try:
            notification = Notification.objects.get(
                id_notification=notification_id, 
                student=request.user.student
            )
            notification.is_read = True
            notification.save()
            return Response({'success': True})
        except Notification.DoesNotExist:
            return Response({'error': 'Уведомление не найдено'}, status=status.HTTP_404_NOT_FOUND)


# ==============================================================================
# ЗАЯВКИ СТУДЕНТОВ ДЛЯ КУРАТОРА
# ==============================================================================

class CuratorStudentRequestsView(generics.ListAPIView):
    """GET: Список заявок студентов группы куратора"""
    serializer_class = StudentRequestSerializer
    permission_classes = [permissions.IsAuthenticated, IsCurator]

    def get_queryset(self):
        # Получаем все группы, где текущий пользователь является куратором
        curator_groups = Group.objects.filter(curator=self.request.user)
        
        # Если пользователь - админ, показываем все заявки
        if self.request.user.is_staff:
            return StudentRequest.objects.all().select_related('student__user', 'student__group').order_by('-created_at')
        
        # Получаем всех студентов этих групп
        students = Student.objects.filter(group__in=curator_groups)
        
        # Возвращаем заявки этих студентов
        return StudentRequest.objects.filter(student__in=students).select_related('student__user', 'student__group').order_by('-created_at')

class CuratorUpdateRequestView(APIView):
    """PATCH: Обновление статуса и комментария заявки студента куратором"""
    permission_classes = [permissions.IsAuthenticated, IsCurator]

    def patch(self, request, request_id):
        try:
            req = StudentRequest.objects.get(id_request=request_id)
            
            # Проверяем, что студент из группы, где текущий пользователь является куратором
            # (или пользователь is_staff)
            if not request.user.is_staff:
                is_curator_of_group = Group.objects.filter(curator=request.user, id=req.student.group.id).exists()
                if not is_curator_of_group:
                    return Response({'error': 'Нет доступа к этой заявке'}, status=status.HTTP_403_FORBIDDEN)
            
            # Обновляем только статус и комментарий
            serializer = StudentRequestSerializer(req, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                
                # Создаем уведомление для студента об ответе
                from .models import Notification
                Notification.objects.create(
                    student=req.student,
                    title=f"Ответ по заявке: {req.get_request_type_display()}",
                    message=f"Ваша заявка переведена в статус: {req.get_status_display()}. Комментарий: {request.data.get('comment', 'Без комментария')}",
                    is_read=False
                )
                
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except StudentRequest.DoesNotExist:
            return Response({'error': 'Заявка не найдена'}, status=status.HTTP_404_NOT_FOUND)


# ==============================================================================
# ПРАКТИКА СТУДЕНТА
# ==============================================================================

class StudentPracticeView(generics.RetrieveAPIView):
    """GET: Информация о практике текущего студента"""
    serializer_class = StudentPracticePlaceSerializer
    permission_classes = [permissions.IsAuthenticated, IsStudent]

    def get_object(self):
        student = self.request.user.student
        # Возвращаем последнее место практики студента
        return StudentPracticePlace.objects.filter(student=student).order_by('-id_place').first()


# ==============================================================================
# ПРАКТИКА СТУДЕНТОВ ДЛЯ КУРАТОРА
# ==============================================================================

class CuratorStudentPracticeView(generics.ListAPIView):
    """GET: Список практики студентов группы куратора"""
    serializer_class = StudentPracticePlaceSerializer
    permission_classes = [permissions.IsAuthenticated, IsCurator]

    def get_queryset(self):
        curator_groups = Group.objects.filter(curator=self.request.user)
        students = Student.objects.filter(group__in=curator_groups)
        return StudentPracticePlace.objects.filter(student__in=students).select_related('student__user', 'organization')


# ==============================================================================
# ПРАКТИКА ДЛЯ ПРЕПОДАВАТЕЛЯ
# ==============================================================================

class TeacherPracticeStudentsView(generics.ListAPIView):
    """GET: Список студентов с информацией о практике для преподавателя"""
    serializer_class = StudentPracticePlaceSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        # Возвращаем все места практики (преподаватель видит всех студентов)
        return StudentPracticePlace.objects.all().select_related('student__user', 'organization')


class TeacherUpdatePracticePlaceView(APIView):
    """PATCH: Обновление места практики студента преподавателем"""
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, place_id):
        try:
            place = StudentPracticePlace.objects.get(id_place=place_id)
            
            # Обновляем только разрешённые поля
            allowed_fields = ['organization', 'position', 'status']
            for field in allowed_fields:
                if field in request.data:
                    setattr(place, field, request.data[field])
            
            place.save()
            
            serializer = StudentPracticePlaceSerializer(place)
            return Response(serializer.data)
        except StudentPracticePlace.DoesNotExist:
            return Response({'error': 'Место практики не найдено'}, status=status.HTTP_404_NOT_FOUND)


class TeacherApproveDiaryEntryView(APIView):
    """PATCH: Одобрение записи в дневнике практики"""
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, entry_id):
        try:
            entry = PracticeDiary.objects.get(id_entry=entry_id)
            entry.is_approved_by_org = True
            entry.save()
            
            serializer = PracticeDiarySerializer(entry)
            return Response(serializer.data)
        except PracticeDiary.DoesNotExist:
            return Response({'error': 'Запись в дневнике не найдена'}, status=status.HTTP_404_NOT_FOUND)
